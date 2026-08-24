from pathlib import Path
from openpyxl import load_workbook

SOURCE_TO_TEMPLATE = {
    "номер ктк": "номер ктк",
    "REF": "Букинг",
    "статус": "Клиент",
    "фактический объем, м3": " объем, м3",
    "вес, кг": "вес, кг",
    "кол-во мест": "кол-во мест",
}


def normalize(value):
    if value is None:
        return ""
    return str(value).strip()


def header_map(ws):
    result = {}
    for cell in ws[1]:
        if cell.value is not None:
            result[normalize(cell.value).lower()] = cell.column
    return result


def find_column(headers, name):
    key = normalize(name).lower()
    if key not in headers:
        raise ValueError(f"В исходном файле не найдена колонка: {name}")
    return headers[key]


def generate_proработка(source_path: str, template_path: str, output_path: str):
    source_wb = load_workbook(source_path, data_only=True)
    source_ws = source_wb.active
    source_headers = header_map(source_ws)

    template_wb = load_workbook(template_path)
    template_ws = template_wb.active
    template_headers = header_map(template_ws)

    # Проверяем все необходимые колонки до начала генерации.
    source_cols = {
        src: find_column(source_headers, src)
        for src in SOURCE_TO_TEMPLATE
    }

    template_cols = {
        dst: find_column(template_headers, dst)
        for dst in SOURCE_TO_TEMPLATE.values()
    }

    # Строка 1 — заголовки. Заполняем строки 2+.
    target_row = 2
    for source_row in range(2, source_ws.max_row + 1):
        # Полностью пустые строки пропускаем.
        values = [
            source_ws.cell(source_row, col).value
            for col in source_cols.values()
        ]
        if all(v in (None, "") for v in values):
            continue

        for src_name, dst_name in SOURCE_TO_TEMPLATE.items():
            src_col = source_cols[src_name]
            dst_col = template_cols[dst_name]
            template_ws.cell(target_row, dst_col).value = source_ws.cell(
                source_row, src_col
            ).value

        target_row += 1

    # Очищаем оставшиеся строки шаблона, если исходных строк меньше.
    for row in range(target_row, template_ws.max_row + 1):
        for col in template_cols.values():
            template_ws.cell(row, col).value = None

    template_wb.save(output_path)
